###############################################################################################
#                                                                                             #
#    /$$$$$$$                               /$$           /$$$$$$$$                 /$$ /$$   #
#   | $$__  $$                             |__/          | $$_____/                |__/| $$   #
#   | $$  \ $$ /$$$$$$   /$$$$$$   /$$$$$$$ /$$  /$$$$$$ | $$    /$$$$$$   /$$$$$$$ /$$| $$   #
#   | $$$$$$$//$$__  $$ /$$__  $$ /$$_____/| $$ /$$__  $$| $$$$$|____  $$ /$$_____/| $$| $$   #
#   | $$____/| $$  \__/| $$$$$$$$| $$      | $$| $$  \ $$| $$__/ /$$$$$$$| $$      | $$| $$   #
#   | $$     | $$      | $$_____/| $$      | $$| $$  | $$| $$   /$$__  $$| $$      | $$| $$   #
#   | $$     | $$      |  $$$$$$$|  $$$$$$$| $$|  $$$$$$/| $$  |  $$$$$$$|  $$$$$$$| $$| $$   #
#   |__/     |__/       \_______/ \_______/|__/ \______/ |__/   \_______/ \_______/|__/|__/   #
#                                                                                             #
#            Buscador de listas de precios de Tienda del Cardan, Bulonera Camba,              #
#             Rosario Agro Industrial y VTM Transmisiones, con actualización                  #
#                                  automática desde internet.                                 #
#                                                                                             #
#                      Autor: Angelo Gallardi (angelogallardi@gmail.com)                      #
#                                                                                             #
###############################################################################################



# -----------------------
# Librerías estándar
# -----------------------
import os
import re
import sqlite3
import subprocess
import sys
import tempfile
import threading
import winreg
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote, urlparse

# -----------------------
# Librerías de terceros
# -----------------------
import bs4
import openpyxl
import requests
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from PyQt6 import uic
from PyQt6.QtCore import (QLibraryInfo, QObject, Qt, QThread, QTranslator,
                          QUrl, pyqtSignal)
from PyQt6.QtGui import QColor, QDesktopServices, QFont, QIcon, QPen
from PyQt6.QtWidgets import (QApplication, QDialog, QHeaderView, QMainWindow,
                             QMessageBox, QStyledItemDelegate,
                             QTableWidgetItem)

# -----------------------
# Módulos del proyecto
# -----------------------
from utils import (APP_DATA_DIR, BRAND_COUNT, CAMBA_SHEETS, CURRENT_VERSION,
                   DB_DIR, DB_PATH, LISTS_DIR, MOST_USED_PRODUCTS_CAMBA,
                   MOST_USED_PRODUCTS_ETMA, MOST_USED_PRODUCTS_HH,
                   MOST_USED_PRODUCTS_VTM, REPO_NAME, REPO_OWNER, ROSARIO_URLS,
                   SETTINGS, get_default_browser_exe)



class UpdateChecker(QObject):
	finished = pyqtSignal(dict) 

	def run(self):
		"""
		Comprueba en GitHub si hay una versión nueva, y emite una diccionario con
		el resultado de la búsqueda.
		"""

		# NOTA: Se movió la verificación de actualizaciones desde MainWindow a un 
		# hilo aparte porque requests.get(url, timeout=3) bloqueaba la interfaz 
		# hasta 3 segundos. Esto provocaba que la ventana no se renderice correctamente
		# mostrando descentrados el QMessageBox de actualización o el QDialog de 
		# progreso al inicio.
		# Ahora se ejecuta en segundo plano para evitar congelamientos.

		url = f'https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/releases/latest'
		
		try:
			# Consulto el último release en GitHub
			response = requests.get(url, timeout=3)
			response.raise_for_status()
			data = response.json()
			latest_version = data['tag_name'].lstrip('v')

			# Comparo versiones
			if latest_version != CURRENT_VERSION:
				download_url = data['assets'][0]['browser_download_url'] # No necesito iterar (un solo asset siempre, el .exe)
				# Hay actualización, emito los datos
				self.finished.emit({
					'has_update': True, 
					'version': latest_version, 
					'url': download_url
				})
				return

		except Exception:
			# Si falla (sin internet o error de API), lo ignoro
			pass 

		# Si llegué acá, no hay actualización o falló la conexión
		self.finished.emit({'has_update': False})


class UpdateDownloader(QObject):
	progress_changed = pyqtSignal(int)
	message_changed = pyqtSignal(str)
	finished = pyqtSignal(str)


	def __init__(self, download_url):
		super().__init__()
		self.download_url = download_url
		self.is_cancelled = False


	def cancel(self):
		"""Activa la bandera para frenar la descarga."""
		self.is_cancelled = True


	def run(self):
		installer_path = Path(os.getenv('TEMP')) / 'PrecioFacil_Update.exe'

		try:
			self.message_changed.emit('Conectando con el servidor...')
			response = requests.get(self.download_url, stream=True, timeout=10)
			response.raise_for_status()

			# Obtengo el tamaño total del archivo para calcular el porcentaje
			total_size = int(response.headers.get('content-length', 0))
			downloaded_size = 0

			self.message_changed.emit('Descargando actualización...')
			with open(installer_path, 'wb') as f:
				for chunk in response.iter_content(chunk_size=8192):
					if self.is_cancelled:
						break # Corta el bucle si el usuario canceló
					if chunk:
						f.write(chunk)
						downloaded_size += len(chunk)
						if total_size > 0:
							# Calculo el porcentaje y emito la señal
							percentage = int((downloaded_size / total_size) * 100)
							self.progress_changed.emit(percentage)

			# Si se canceló, limpio y salgo
			if self.is_cancelled:
				if installer_path.exists():
					installer_path.unlink() # Borro la basura
				self.finished.emit('cancelled') # Aviso que fue cancelado
				return

			# Si llegué acá, la descarga terminó bien
			self.message_changed.emit('Iniciando instalador...')
			subprocess.Popen([installer_path, '/SILENT']) # Instalación silenciosa (solo barra de progreso)
			self.finished.emit('success') # Aviso que fue exitoso

		except Exception as e:
			# Si hay error (ej. se corta internet), limpio y aviso
			if installer_path.exists():
				try:
					installer_path.unlink()
				except OSError:
					pass # Si Windows lo tiene bloqueado por alguna razón, lo ignoramos

			self.finished.emit(f'error|{str(e)}') # Aviso que hubo error


class DataProcessor(QObject):
	progress_changed = pyqtSignal(int)
	message_changed = pyqtSignal(str)
	finished = pyqtSignal(dict) # Emite un diccionario con productos + reporte


	def __init__(self):
		super().__init__()
		# Sesión persistente (acelera mucho, evita abrir una conexión nueva cada vez)
		self.session = requests.Session()

		# Variable para almacenar el token de la sesión
		self.tdc_token = None


	def update_progress(self, points_to_add, message=None):
		"""Suma puntos al progreso total y actualiza la UI."""

		self.current_progress += points_to_add
		
		# Evito pasarme de 100 por si hay algún redondeo raro
		if self.current_progress > 100:
			self.current_progress = 100
			
		if message:
			self.message_changed.emit(message)
		
		# Emito el entero a la barra de progreso
		self.progress_changed.emit(int(self.current_progress))


	# CÓDIGO PRINCIPAL
	# ------------------------------------------------------------------------------------------

	def run(self):
		"""
		Método principal ejecutado por el hilo secundario para gestionar la descarga
		y procesamiento de las listas.
		"""

		# Inicializo variables
		self.current_progress = 0
		self.report = {}
		self.all_data = {
			'hh': {'products': [], 'date': ''},
			'etma': {'products': [], 'date': ''},
			'camba': {'products': [], 'date': ''},
			'vtm': {'products': [], 'date': ''},
			'report': self.report
		}

		# Calculo puntajes de progreso
		camba_files = 1 + len(CAMBA_SHEETS) # 1 excel + N pdfs
		rosario_files = len(ROSARIO_URLS)
		self.points_per_brand = 100 // BRAND_COUNT
		self.points_per_file_camba = self.points_per_brand / camba_files
		self.points_per_file_rosario = self.points_per_brand / rosario_files

		self.update_progress(0, 'Iniciando carga...')

		# 1. PROCESO TIENDA DEL CARDAN (HH y ETMA) VÍA API PROTEGIDA
		self.update_progress(0, 'Conectando con Tienda del Cardan...')
		if self.login_tienda_cardan():
			for brand in ('hh', 'etma'):
				self.update_progress(0, f'Procesando {brand.upper()}...')
				self.process_tdc_brand(brand)
		else:
			self.handle_supplier_down(('hh', 'etma'), 'login_failed')

		# 2. PROCESO CAMBA VÍA WEB SCRAPING
		supplier_url_camba = self.get_url_from_settings('camba')
		if not supplier_url_camba:
			self.handle_supplier_down(('camba',), 'no_url')
		else:
			# Obtengo el HTML de la URL, lo parseo, y proceso excel y pdfs
			try:
				html = self.download_html(supplier_url_camba)
				soup = bs4.BeautifulSoup(html, 'html.parser')
				self.update_progress(0, 'Procesando CAMBA...')
				self.process_camba_brand(soup)
				self.process_camba_pdfs(soup)
			except Exception:
				self.handle_supplier_down(('camba',), 'no_access')

		# 3. PROCESO ROSARIO AGRO
		self.update_progress(0, 'Procesando ROSARIO AGRO...')
		self.process_rosario_pdfs()

		# 4. PROCESO VTM
		self.update_progress(0, 'Procesando VTM...')
		self.process_vtm()

		self.update_progress(100, '¡Carga completada!')
		self.finished.emit(self.all_data) # Devuelvo datos recolectados al MainWindow


	# PROCESAMIENTO POR MARCA
	# ------------------------------------------------------------------------------------------

	def login_tienda_cardan(self):
		"""Inicia sesión en Firebase y guarda el token para usarlo en las descargas."""

		
		# 1. Cargo las variables desde el archivo .env
		# -------------------------------------------------------

		load_dotenv()

		# Obtengo las credenciales
		api_key = os.getenv('FIREBASE_API_KEY')
		email = os.getenv('USER_EMAIL')
		password = os.getenv('USER_PASSWORD')

		# Valido que las variables existan
		if not api_key or not email or not password:
			return False # Falta el .env o credenciales en el mismo

		# 2. Configuración de Firebase
		# -------------------------------------------------------

		# Endpoint oficial de Firebase para login con email/password
		url_firebase = f'https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={api_key}'

		credenciales = {
			'email': email, 
			'password': password, 
			'returnSecureToken': True # Obligatorio para que me devuelva el token
		}

		# 3. Inicio sesión en Firebase
		# -------------------------------------------------------

		try:
			response = self.session.post(url_firebase, json=credenciales, timeout=10)
			response.raise_for_status()
			self.tdc_token = response.json().get('idToken') # Extraigo el token de la respuesta JSON
			return True
		except Exception:
			return False


	def process_tdc_brand(self, brand):
		"""Procesa las marcas de TDC directamente desde la API."""

		step_points = self.points_per_brand / 2 # 1=descargar, 2=procesar (link ya lo tengo)

		# Link de la lista
		list_url = f'https://app.tiendadecardan.com.ar/api/precios/descargar-base?marca={brand.upper()}&format=xlsx'

		# Descargo la lista
		try:
			excel_file_path = self.download_excel_file(list_url, brand)
			self.update_progress(step_points)
		except Exception:
			self.check_local_excel_list(brand, 'no_download')
			self.update_progress(step_points) # Sumo el paso restante (procesar)
			return

		# Proceso excel descargado
		try:
			self.process_excel(excel_file_path, brand)
		except Exception:
			self.report.setdefault(brand, {})['excel'] = {
				'local_status': 'local_error'
			}

		self.update_progress(step_points)


	def process_camba_brand(self, soup):
		"""Busca la URL de la lista excel de CAMBA en el soup, la descarga y la procesa."""

		brand = 'camba'
		step_points = self.points_per_file_camba / 3 # 1=link, 2=descargar, 3=procesar

		# Busco link de la lista
		list_url = self.get_camba_list_url_from_soup(soup)
		self.update_progress(step_points)
		if not list_url:
			self.check_local_excel_list(brand, 'no_link')
			self.update_progress(step_points * 2) # Como fui al fallback, sumo de golpe los 2 pasos restantes (descargar y procesar)
			return

		# Descargo la lista
		try:
			excel_file_path = self.download_excel_file(list_url, brand)
			camba_last_date = self.resolve_camba_date(soup, excel_file_path) # Porque CAMBA no tiene la fecha en el excel
			SETTINGS.setValue('camba_last_date', camba_last_date)
			self.update_progress(step_points)
		except Exception:
			self.check_local_excel_list(brand, 'no_download')
			self.update_progress(step_points) # Solo sumo el paso restante (procesar)
			return

		# Proceso excel descargado
		try:
			self.process_excel(excel_file_path, brand)
		except Exception:
			self.report.setdefault(brand, {})['excel'] = {
				'local_status': 'local_error'
			}

		self.update_progress(step_points)


	def process_camba_pdfs(self, soup):
		"""Inicia la descarga paralela de los PDFs de CAMBA encontrados en el soup."""

		# Construyo ruta de la carpeta destino
		base_path = LISTS_DIR / 'camba'
		base_path.mkdir(parents=True, exist_ok=True)

		with ThreadPoolExecutor(max_workers=5) as executor:
			for sheet_num in CAMBA_SHEETS:
				executor.submit(
					self.download_camba_pdf,
					sheet_num,
					soup,
					base_path
				)


	def download_camba_pdf(self, sheet_num, soup, base_path):
		"""Descarga un PDF específico de CAMBA según el número de hoja."""

		# Busco el link de la hoja
		a_elem = soup.find(
			'a',
			href=True,
			string=lambda s: s and f'Hoja {sheet_num}' in s
		)
		if not a_elem:
			self.check_local_pdf_list('camba', sheet_num, 'no_link')
			self.update_progress(self.points_per_file_camba) # Sumo antes de retornar
			return

		# Obtengo la ruta completa
		pdf_url = a_elem['href']
		pdf_original_name = pdf_url.split('=')[-1] + '.pdf'
		pdf_file_path = base_path / pdf_original_name

		# Si ya existe con este nombre exacto, lo salteamos
		if pdf_file_path.exists():
			self.update_progress(self.points_per_file_camba) # Sumo antes de retornar
			return

		# Descargo el PDF
		try:
			response = self.session.get(pdf_url, timeout=10)
			response.raise_for_status()

			# Borro versiones viejas del mismo número de hoja
			for old_pdf in base_path.glob(f'Hoja{sheet_num}*.pdf'):
				old_pdf.unlink()

			# Guardo el archivo descargado
			with open(pdf_file_path, 'wb') as f:
				f.write(response.content)

		except Exception:
			self.check_local_pdf_list('camba', sheet_num, 'no_download')

		# Sumo al final si todo el proceso normal terminó
		self.update_progress(self.points_per_file_camba)


	def process_rosario_pdfs(self):
		"""Inicia la descarga paralela de los PDFs de ROSARIO AGRO."""

		base_path = LISTS_DIR / 'rosario'
		base_path.mkdir(parents=True, exist_ok=True)

		with ThreadPoolExecutor(max_workers=5) as executor:
			for pdf_url in ROSARIO_URLS:
				executor.submit(
					self.download_rosario_pdf,
					pdf_url,
					base_path
				)


	def download_rosario_pdf(self, pdf_url, base_path):
		"""Descarga un PDF de ROSARIO AGRO desde la URL indicada, sobreescribiendo."""

		# Obtengo la ruta completa
		pdf_original_name = pdf_url.split('=')[-1]
		pdf_file_path = base_path / pdf_original_name

		# Descargo el PDF
		try:
			response = self.session.get(pdf_url, timeout=10)
			response.raise_for_status()

			with open(pdf_file_path, 'wb') as f:
				f.write(response.content)

		except Exception:
			self.check_local_pdf_list('rosario', pdf_file_path.stem, 'no_download') # paso solo nombre del PDF
		
		self.update_progress(self.points_per_file_rosario)


	def process_vtm(self):
		"""Descarga y procesa la lista de VTM usando Playwright en segundo plano."""
		
		brand = 'vtm'
		step_points = self.points_per_brand / 2 # 1=descargar, 2=procesar
		
		base_path = LISTS_DIR / brand
		base_path.mkdir(parents=True, exist_ok=True)
		excel_file_path = None

		try:
			from playwright.sync_api import sync_playwright

			with sync_playwright() as p:
				launch_args = {'headless': True}
				
				# Busco el navegador predeterminado para no depender del Chromium interno
				browser_exe = get_default_browser_exe()
				if browser_exe:
					launch_args['executable_path'] = browser_exe
				else:
					# Fallback seguro para Windows: usa Edge si no encuentra el predeterminado
					launch_args['channel'] = 'msedge' 

				browser = p.chromium.launch(**launch_args)
				context = browser.new_context(accept_downloads=True)
				page = context.new_page()

				page.goto('https://vtm-lista.pages.dev/', wait_until='networkidle')
				page.wait_for_timeout(2000)

				# 1. Abro el combobox "Exportar Excel"
				toggle = page.locator('button.toolbar-btn-export')
				toggle.wait_for(state='visible', timeout=10000)
				toggle.click()

				# Espero a que el menú quede realmente desplegado (aria-expanded="true")
				page.wait_for_selector("button.toolbar-btn-export[aria-expanded='true']", timeout=5000)
				page.wait_for_timeout(300) # pequeño margen para animación/render del menú

				# 2. Click en el ítem "Todo el catálogo" dentro del menú
				boton = page.get_by_role('menuitem', name='Todo el catálogo')
				boton.wait_for(state='visible', timeout=10000)

				with page.expect_download() as download_info:
					boton.click()
				
				download = download_info.value
				
				# Obtengo el nombre original y armo la ruta final
				original_filename = download.suggested_filename
				excel_file_path = base_path / original_filename
				
				# Borro excel previo de la carpeta
				for old_excel in base_path.glob('*.xlsx'):
					old_excel.unlink()

				# Guardo el nuevo archivo con su nombre original
				download.save_as(excel_file_path)
				browser.close()
				
			self.update_progress(step_points)
			
		except Exception as e:
			# print(f'Error al descargar VTM: {e}')
			self.check_local_excel_list(brand, 'no_download')
			self.update_progress(step_points)
			return

		# Proceso excel descargado
		try:
			self.process_excel(excel_file_path, brand)
		except Exception:
			self.report.setdefault(brand, {})['excel'] = {
				'local_status': 'local_error'
			}

		self.update_progress(step_points)


	# FALLBACKS (hubo error y se deben buscar listas locales descargadas previamente)
	# ------------------------------------------------------------------------------------------

	def handle_supplier_down(self, brands, reason):
		"""
		Fallback general llamado cuando falla la conexión inicial con un proveedor:
		  * Falla el inicio de sesión en la API (TDC).
		  * No hay URL configurada o no se pudo obtener el HTML (CAMBA).
		Llama al fallback de marca por cada marca del proveedor. Además, 
		si es CAMBA, chequea las hojas PDF locales.
		"""

		for brand in brands:
			self.update_progress(0, f'Procesando {brand.upper()}...')
			self.check_local_excel_list(brand, reason)

			# Si es CAMBA, compruebo PDFs locales
			if brand == 'camba':
				self.update_progress(self.points_per_file_camba) # excel recién procesado arriba
				for sheet_num in CAMBA_SHEETS:
					self.check_local_pdf_list('camba', sheet_num, reason)
					self.update_progress(self.points_per_file_camba)
			else:
				self.update_progress(self.points_per_brand)


	def check_local_excel_list(self, brand, reason):
		"""
		Fallback por marca llamado cuando falla la obtención del archivo original:
		  * No se encontró el link de la lista (CAMBA).
		  * Falla la petición de descarga (TDC / CAMBA).
		Comprueba si existe un excel local previamente descargado y lo procesa.
		"""

		base_path = LISTS_DIR / brand
		excel_file_path = None

		if base_path.exists():
			excel_files = list(base_path.glob('*.xlsx'))
			if excel_files:
				excel_file_path = excel_files[0]

		if excel_file_path:
			try:
				self.process_excel(excel_file_path, brand)
				local_status = 'local_used'
			except Exception:
				local_status = 'local_error'
		else:
			local_status = 'local_missing'

		self.report.setdefault(brand, {})['excel'] = {
			'reason': reason,
			'local_status': local_status
		}


	def check_local_pdf_list(self, brand, identifier, reason):
		"""
		Verifica existencia local de un PDF cuando falla descarga/encontrar link.
		* Para CAMBA: identifier es el número de hoja (por ej: '02')
		* Para ROSARIO: identifier es el nombre del archivo (ej: 'Cuchillas_Jardin')
		"""

		base_path = LISTS_DIR / brand
		has_local = False

		if base_path.exists():
			if brand == 'camba':
				# Busco PDFs con el número de hoja
				pdf_files = list(base_path.glob(f'Hoja{identifier}*.pdf'))
				if pdf_files:
					has_local = True
			elif brand == 'rosario':
				if (base_path / f'{identifier}.pdf').exists():
					has_local = True

		# Guardo resultado para agruparlo después
		pdfs = self.report.setdefault(brand, {}).setdefault('pdfs', {})
		entry = pdfs.setdefault(reason, {'missing': [], 'local': []})
		if has_local:
			entry['local'].append(identifier)
		else:
			entry['missing'].append(identifier)


	#  AUXILIARES
	# ------------------------------------------------------------------------------------------

	def get_url_from_settings(self, supplier):
		return SETTINGS.value(f'supplier_urls/{supplier}', '', type=str)


	def download_html(self, url):
		response = requests.get(url, timeout=10)
		response.raise_for_status()
		return response.text


	def get_camba_list_url_from_soup(self, soup):
		"""Obtiene del sitio de CAMBA el link actual de la lista formato sábana."""

		# Busco el título correcto
		h2_elem = soup.find(
			'h2', 
			string=lambda s: s and 'Lista de precios formato sabana' in s.strip()
		)
		if not h2_elem:
			return None

		# Busco el link
		a_elem = h2_elem.find_parent('a')

		return a_elem['href'] if a_elem else None


	def download_excel_file(self, url, brand):
		"""Descarga el excel en la carpeta correspondiente."""

		# Construyo ruta de la carpeta destino
		base_path = LISTS_DIR / brand
		base_path.mkdir(parents=True, exist_ok=True)

		# 1. Configuro los encabezados dinámicamente según la marca
		headers = {
			'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
		}

		# Si es TDC, le inyecto el token
		if brand in ('hh', 'etma') and self.tdc_token:
			headers['Authorization'] = f'Bearer {self.tdc_token}'

		# 2. Descargo el archivo
		response = self.session.get(url, headers=headers, timeout=10)
		response.raise_for_status()

		# Borro excel previo
		for old_excel_file in base_path.glob('*.xlsx'):
			old_excel_file.unlink()

		# 3. Guardo el archivo dependiendo de si es Excel directo o ZIP
		if brand in ('hh', 'etma'):
			# Busco el nombre original en los encabezados del servidor
			excel_original_name = f'lista_precios_{brand}.xlsx' # Nombre por defecto (Fallback)

			content_disposition = response.headers.get('content-disposition')
			if content_disposition and 'filename=' in content_disposition:
				# Extraigo el texto que está después de "filename=" y limpio comillas
				# Ejemplo: attachment; filename="lista_hh_08_2026.xlsx" -> lista_hh_08_2026.xlsx
				excel_original_name = content_disposition.split('filename=')[1].split(';')[0].strip('"\'')

			# Armo la ruta con el nombre detectado
			excel_file_path = base_path / excel_original_name		

			# Guardo el archivo descargado
			with open(excel_file_path, 'wb') as f:
				f.write(response.content)

		else: # la URL entrega un zip (Camba)
			with zipfile.ZipFile(BytesIO(response.content)) as z:
				for name in z.namelist():
					if name.lower().endswith('.xlsx'):
						excel_file_path = base_path / name
						with z.open(name) as source, open(excel_file_path, 'wb') as target:
							target.write(source.read())
						break

		return excel_file_path


	def resolve_camba_date(self, soup, excel_file_path):
		"""
		Determina qué fecha asociar al excel descargado de CAMBA usando, por prioridad:
		1) HTML
		2) nombre del archivo
		3) fecha actual
		"""

		# Busco en el HTML
		date = self.extract_camba_date_from_soup(soup)
		if date:
			return date

		# Busco en el nombre de archivo
		date = self.extract_date_from_filename(excel_file_path)
		if date:
			return date

		# Como último recurso: fecha actual
		return datetime.now().strftime('%d/%m/%Y')


	def extract_camba_date_from_soup(self, soup):
		"""Extrae la fecha actual de las listas de CAMBA desde el HTML."""

		a_elem = soup.find(
			'a',
			href=True,
			string=lambda s: s and 'lista indice' in s.strip().lower()
		)

		if not a_elem:
			return None

		match = re.search(r'\d{2}/\d{2}/\d{4}', a_elem.get_text())
		return match.group() if match else None


	def extract_date_from_filename(self, path):
		"""Extrae la fecha desde el nombre del archivo excel de CAMBA."""
		
		match = re.search(r'\d{2}[-_]\d{2}[-_]\d{4}', path)
		if match:
			return re.sub(r'[-_]', '/', match.group())
		return None


	def process_excel(self, excel_file_path, brand):
		"""
		Lee el excel, extrae la fecha de validez y los productos, y guarda todo 
		en all_data.
		"""

		# Creo workbook y extraigo la hoja de productos
		wb = openpyxl.load_workbook(excel_file_path)
		sheet = wb[wb.sheetnames[0]]

		# Busco letras de columnas de producto
		header_cols = self.search_header_cols(sheet, brand)

		# Busco número de fila de primer producto
		first_row = self.search_first_row(sheet, header_cols['price_col'])

		# Extraigo la fecha de validez de precios
		validity_date = self.extract_validity_date(brand, sheet)

		# Extraigo todos los productos en una lista de diccionarios
		products = self.obtain_products(sheet, first_row, header_cols, brand)

		# Guardo los datos recolectados en el diccionario general
		self.all_data[brand]['products'] = products
		self.all_data[brand]['date'] = validity_date


	def search_header_cols(self, sheet, brand):
		"""Retorna un diccionario con la posición (letra) de cada columna."""

		default_cols = {
			'hh': {
				'code_col': 'A',
				'subcategory_col': None,
				'description_col': 'B',
				'price_col': 'C'
			},
			'etma': {
				'code_col': 'A',
				'subcategory_col': None,
				'description_col': 'B',
				'price_col': 'C'
			},
			'camba': {
				'code_col': 'B',
				'subcategory_col': 'J',
				'description_col': 'C',
				'price_col': 'E'
			},
			'vtm': {
				'code_col': 'A',
				'subcategory_col': 'F',
				'description_col': 'C',
				'price_col': 'D'
			}
		}

		# Valores por defecto en caso de que no encuentre alguna
		code_col = default_cols[brand]['code_col']
		subcategory_col = default_cols[brand]['subcategory_col']
		description_col = default_cols[brand]['description_col']
		price_col = default_cols[brand]['price_col']

		header_row = None

		# Busco la fila de encabezados
		for row in sheet['A1':'F20']:
			for cell in row:
				value = str(cell.value or '').strip().lower()

				if value in ('sku', 'referencia interna', 'codigo vtm'):
					header_row = cell.row
					code_col = get_column_letter(cell.column)
					break

			if header_row is not None:
				break

		# Si encontré la fila, identifico las demás columnas
		if header_row is not None:
			for header_cell in sheet[header_row]:
				value = str(header_cell.value or '').strip().lower()

				if 'categoría' in value or 'categoria' in value or 'rubro' in value:
					subcategory_col = get_column_letter(header_cell.column)
				elif 'nombre' in value or 'producto' in value or 'descripcion' in value or 'descripción' in value:
					description_col = get_column_letter(header_cell.column)
				elif 'precio' in value:
					price_col = get_column_letter(header_cell.column)

		return {
			'code_col': code_col,
			'subcategory_col': subcategory_col,
			'description_col': description_col,
			'price_col': price_col
		}


	def search_first_row(self, sheet, price_col):
		"""Retorna la fila donde comienzan los productos."""

		for cell in sheet[price_col]:
				value = cell.value

				# Evito trabajo innecesario (evito convertir None)
				if value is None:
					continue

				# Compruebo si es un monto
				try:
					float(str(value).replace('.', '').replace(',', '.'))
					return cell.row
				except (ValueError, TypeError):
					continue


	def extract_validity_date(self, brand, sheet):
		"""Busca y retorna la fecha de validez del excel de la marca."""

		# Para CAMBA se busca en la configuración guardada
		if brand == 'camba':
			stored_date = SETTINGS.value('camba_last_date', '', type=str)
			
			if stored_date:
				return f'📆 Precios válidos para el: {stored_date}'

			return '📆 Fecha no disponible'

		# VTM tiene la fecha en las últimas filas
		if brand == 'vtm':
			max_r = sheet.max_row
			min_r = max(1, max_r - 20) # Reviso solo las últimas 20 filas
			for row in sheet.iter_rows(min_row=min_r, max_row=max_r):
				for cell in row:
					value = str(cell.value or '').strip().lower()
					if 'generado:' in value:
						match = re.search(r'\d{1,2}[-/]\d{1,2}[-/]\d{4}', value)
						if match:
							date = match.group().replace('-', '/')
							return f'📆 Precios válidos para el: {date}'
			return '📆 Fecha no encontrada'

		# Para HH o ETMA se busca en las primeras celdas
		for row in sheet['A1':'F20']:
			for cell in row:
				# Evito trabajo innecesario (no analizo celdas vacías)
				if not cell.value:
					continue

				value = str(cell.value).strip().lower()

				# La celda de fecha contiene "Fecha de actualizacion"
				if 'fecha' not in value:
					continue

				match = re.search(r'\d{2}[-/]\d{2}[-/]\d{4}', value)
				if match:
					date = match.group().replace('-', '/')
					return f'📆 Precios válidos para el: {date}'

		return '📆 Fecha no encontrada'


	def obtain_products(self, sheet, first_row, header_cols, brand):
		"""Crea lista de diccionarios de productos para filtrar."""

		products = []
		for row in range(first_row, sheet.max_row + 1):
			if self.is_valid_row(sheet, row, header_cols):
				# Formateo precios de tipo float o int (necesario para VTM y CAMBA)
				price = sheet[header_cols['price_col'] + str(row)].value

				if isinstance(price, (float, int)):
					price = f'{price:,.2f}'.replace('.', '_').replace(',', '.').replace('_', ',')

				# Creo diccionario sin subcategoría
				product = {
					'code': str(sheet[header_cols['code_col'] + str(row)].value).strip(),
					'description': str(sheet[header_cols['description_col'] + str(row)].value).strip(),
					'price': f'$ {price}'
				}

				# Si es CAMBA o VTM, agrego la subcategoría
				if brand in ('camba', 'vtm') and header_cols['subcategory_col']:
					raw_subcat = str(sheet[header_cols['subcategory_col'] + str(row)].value).strip()

					# Solo corto el prefijo si la marca es Camba
					if brand == 'camba' and ' - ' in raw_subcat:
						product['subcategory'] = raw_subcat.split(' - ', 1)[1].strip()
					else:
						product['subcategory'] = raw_subcat

				products.append(product)

		return products


	def is_valid_row(self, sheet, row, header_cols):
		"""Retorna si una fila corresponde o no a un producto."""

		for col in header_cols.values():
			# Ignoro las columnas que no existen para esta marca (ej: subcategoría en HH/ETMA)
			if col is None:
				continue

			val = sheet[col + str(row)].value

			# Verifico que no sea None y que no sea un string vacío o con espacios
			if val is None or str(val).strip() == '':
				return False

		return True



class ThickGridDelegate(QStyledItemDelegate):
	"""
	Delegado de tabla que superpone un trazo más grueso en los bordes de filas y
	columnas específicas, preservando la apariencia base.
	"""

	def __init__(self, target_rows=(), target_cols=(), parent=None):
		super().__init__(parent)
		self.target_rows = target_rows
		self.target_cols = target_cols


	def set_line_color(self, hex_color):
		"""Actualiza el color dinámicamente cuando cambia el tema."""
		self.line_color = QColor(hex_color)


	def paint(self, painter, option, index):
		# 1. Dibujo el contenido normal de la celda primero
		super().paint(painter, option, index)

		is_target_row = index.row() in self.target_rows
		is_target_col = index.column() in self.target_cols

		if is_target_row or is_target_col:
			painter.save()
			
			# Configuro el color y grosor de la línea
			pen = QPen(self.line_color)
			pen.setWidth(2)
			painter.setPen(pen)

			# Dibujo la línea
			rect = option.rect
			if is_target_row:
				painter.drawLine(rect.bottomLeft(), rect.bottomRight())
			if is_target_col:
				painter.drawLine(rect.topRight(), rect.bottomRight())
				
			painter.restore()



class MainWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		# Cargo la UI
		uic.loadUi('ui/app.ui', self)

		# Título de la ventaba
		self.setWindowTitle(f'PrecioFacil {CURRENT_VERSION}')

		# Establezco la pestaña de Camba como la inicial
		self.tabWidget.setCurrentIndex(3)

		# Inicializo base de datos y tabla de ventiladores
		self.init_db()
		self.load_fans_data()
		self.tableWidget_fans.setItemDelegate(
			ThickGridDelegate(
				target_rows=(8,), 
				parent=self.tableWidget_fans
			)
		)

		# Señales de pushbuttons inferiores
		self.pushButton_theme.clicked.connect(self.change_theme)
		self.pushButton_config.clicked.connect(self.open_config)
		self.pushButton_about.clicked.connect(self.open_about)

		# Señales de pushbuttons de BULONERA CAMBA
		self.pushButton_alemite.clicked.connect(lambda: self.open_pdf('camba', '22', 2))
		self.pushButton_seeger.clicked.connect(lambda: self.open_pdf('camba', '35', 2))
		self.pushButton_arandela_grower.clicked.connect(lambda: self.open_pdf('camba', '10', 4))
		self.pushButton_arandela_plana.clicked.connect(lambda: self.open_pdf('camba', '16', 1))
		self.pushButton_bulon_unc.clicked.connect(lambda: self.open_pdf('camba', '02', 1))
		self.pushButton_bulon_unf.clicked.connect(lambda: self.open_pdf('camba', '07', 1))
		self.pushButton_chaveta_partida.clicked.connect(lambda: self.open_pdf('camba', '19', 1))
		self.pushButton_espina_elastica.clicked.connect(lambda: self.open_pdf('camba', '35', 1))
		self.pushButton_prisionero_cilindrica.clicked.connect(lambda: self.open_pdf('camba', '14', 2))
		self.pushButton_prisionero_sin.clicked.connect(lambda: self.open_pdf('camba', '14', 3))
		self.pushButton_prisionero_cuadrada.clicked.connect(lambda: self.open_pdf('camba', '13', 1))
		self.pushButton_tuerca_exagonal.clicked.connect(lambda: self.open_pdf('camba', '04', 1))
		self.pushButton_tuerca_castillo.clicked.connect(
			lambda: [
				self.open_pdf('camba', '04', 5),
				self.open_pdf('camba', '23', 1)
			]
		)
		self.pushButton_tuerca_torneada.clicked.connect(lambda: self.open_pdf('camba', '23', 1))
		self.pushButton_varilla_camba.clicked.connect(
			lambda: [
				self.open_pdf('camba', '11', 2),
				self.open_pdf('camba', '17', 1)
			]
		)
		self.pushButton_tornillo_metrico.clicked.connect(lambda: self.open_pdf('camba', '13', 2))
		self.pushButton_tornillo_inox.clicked.connect(lambda: self.open_pdf('camba', '36', 8))

		# Señales de pushbuttons de ROSARIO AGRO
		self.pushButton_gummi.clicked.connect(lambda: self.open_pdf('rosario', 'GUMMI'))
		self.pushButton_tupac.clicked.connect(lambda: self.open_pdf('rosario', 'Tupac'))
		self.pushButton_cadena.clicked.connect(lambda: self.open_pdf('rosario', 'Cadenas_LinkBelt'))
		self.pushButton_cruceta.clicked.connect(lambda: self.open_pdf('rosario', 'Crucetas_ETMA'))
		self.pushButton_cuchilla.clicked.connect(lambda: self.open_pdf('rosario', 'Cuchillas_Agro'))
		self.pushButton_forro.clicked.connect(lambda: self.open_pdf('rosario', 'FORRO_DE_EMBRAGUE'))
		self.pushButton_polea.clicked.connect(lambda: self.open_pdf('rosario', 'PoleasHF'))
		self.pushButton_cardan.clicked.connect(lambda: self.open_pdf('rosario', 'Repuestos_cardanicos'))
		self.pushButton_rotula.clicked.connect(lambda: self.open_pdf('rosario', 'Rotulas'))
		self.pushButton_varilla_rosario.clicked.connect(lambda: self.open_pdf('rosario', 'ROSCAS_ACME'))
		self.pushButton_soporte.clicked.connect(lambda: self.open_pdf('rosario', 'Soportes_FKD'))
		self.pushButton_termo.clicked.connect(lambda: self.open_pdf('rosario', 'Termoplasticos'))

		# Señales de comboboxes
		self.comboBox_most_used_hh.activated.connect(self.load_category)
		self.comboBox_most_used_etma.activated.connect(self.load_category)
		self.comboBox_most_used_camba.activated.connect(self.load_category)
		self.comboBox_most_used_vtm.activated.connect(self.load_category)

		# Señales de lineedits
		self.lineEdit_search_hh.textEdited.connect(self.filter_products)
		self.lineEdit_search_etma.textEdited.connect(self.filter_products)
		self.lineEdit_search_camba.textEdited.connect(self.filter_products)
		self.lineEdit_search_vtm.textEdited.connect(self.filter_products)

		# Señales de tablewidgets
		self.tableWidget_search_hh.itemDoubleClicked.connect(self.open_tdc_catalog)
		self.tableWidget_defaults_hh.itemDoubleClicked.connect(self.open_tdc_catalog)
		self.tableWidget_search_etma.itemDoubleClicked.connect(self.open_tdc_catalog)
		self.tableWidget_defaults_etma.itemDoubleClicked.connect(self.open_tdc_catalog)
		self.tableWidget_search_vtm.itemDoubleClicked.connect(self.open_vtm_catalog)
		self.tableWidget_defaults_vtm.itemDoubleClicked.connect(self.open_vtm_catalog)
		self.tableWidget_fans.itemChanged.connect(self.save_fan_cell)

		# Configuraciones visuales varias
		self.format_headers() # Configuro headers de tablas
		self.initialize_theme()
		self.showMaximized() # Ventana maximizada

		# Comienzo comprobando actualizaciones
		self.start_update_check()


	def start_update_check(self):
		"""Inicia la comprobación de actualizaciones en un hilo secundario."""

		# Configuro el hilo y el worker
		self.checker_thread = QThread()
		self.checker_worker = UpdateChecker()
		self.checker_worker.moveToThread(self.checker_thread)

		# Conecto señales de inicio y fin
		self.checker_thread.started.connect(self.checker_worker.run)
		self.checker_worker.finished.connect(self.on_update_check_finished)

		# Limpieza de memoria
		self.checker_worker.finished.connect(self.checker_thread.quit)
		self.checker_worker.finished.connect(self.checker_worker.deleteLater)
		self.checker_thread.finished.connect(self.checker_thread.deleteLater)

		# Arranco el hilo
		self.checker_thread.start()


	def on_update_check_finished(self, result):
		"""Recibe el resultado de GitHub y decide qué hacer."""
		
		# Si el worker detectó una actualización
		if result['has_update']:
			# Pregunto al usuario
			reply = QMessageBox.question(
				self,
				'Actualización disponible',
				f'Hay una nueva versión de PrecioFacil (v{result["version"]}).\n¿Querés actualizar ahora?',
				QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
			)

			if reply == QMessageBox.StandardButton.Yes:
				self.start_update_download(result['url'])
				return # Salgo para no iniciar la carga de datos

		# Si NO hay actualización, o hubo error, o el usuario dijo que NO:
		# Inicio el flujo normal de la aplicación
		self.start_data_processing()


	def start_update_download(self, download_url):
		"""Inicia la descarga de la actualización en un hilo secundario."""

		# Creo el dialog de progreso
		self.downloader_dialog = ProgressDialog('Progreso de la descarga', True, self)

		# Configuro el hilo y el worker
		self.downloader_thread = QThread()
		self.downloader_worker = UpdateDownloader(download_url)
		self.downloader_worker.moveToThread(self.downloader_thread)
		
		# Conecto las señales del worker al dialog
		self.downloader_worker.message_changed.connect(self.downloader_dialog.label.setText)
		self.downloader_worker.progress_changed.connect(self.downloader_dialog.progressBar.setValue)

		# Conecto señal de cancelación de descarga
		# Qt.ConnectionType.DirectConnection obliga a que el método cancel() se ejecute
		# en el momento exacto en que se hace clic
		self.downloader_dialog.rejected.connect(self.downloader_worker.cancel, Qt.ConnectionType.DirectConnection)

		# Conecto señales de ciclo de vida e inicio
		self.downloader_thread.started.connect(self.downloader_worker.run)
		self.downloader_worker.finished.connect(self.on_update_finished)

		# Limpieza de memoria al terminar
		self.downloader_worker.finished.connect(self.downloader_thread.quit)
		self.downloader_worker.finished.connect(self.downloader_worker.deleteLater)
		self.downloader_thread.finished.connect(self.downloader_thread.deleteLater)

		# Inicio el hilo y muestro el dialog de forma modal
		self.downloader_thread.start()
		self.downloader_dialog.exec()


	def on_update_finished(self, status):
		"""Procesa el resultado de la descarga y continúa el flujo."""
		
		# Cierro el dialog
		self.downloader_dialog.accept()

		if status == 'success':
			# Se descargó y se lanzó el instalador, cierro la app
			QApplication.instance().quit()

		elif status.startswith('error|'):
			error_msg = status.split('|')[1]
			QMessageBox.warning(
				self, 
				'Error', 
				f'Se interrumpió la descarga.\nDetalle: {error_msg}'
			)

			# Recién cuando el usuario aprieta Aceptar, la app sigue
			self.start_data_processing()

		elif status == 'cancelled':
			# El usuario lo canceló a mano. No muestro error, arranca normal directo
			self.start_data_processing()


	def start_data_processing(self):
		"""Inicia el procesamiento de las listas de precios en un hilo secundario."""

		# Vacio todo por si es una recarga
		self.empty_everything()

		# Creo el dialog de progreso
		self.processor_dialog = ProgressDialog('Progreso de la carga', False, self)
		
		# Configuro el hilo y el worker
		self.processor_thread = QThread()
		self.processor_worker = DataProcessor()
		self.processor_worker.moveToThread(self.processor_thread)

		# Conecto las señales del worker al dialog
		self.processor_worker.message_changed.connect(self.processor_dialog.label.setText)
		self.processor_worker.progress_changed.connect(self.processor_dialog.progressBar.setValue)
		
		# Conecto señales de ciclo de vida e inicio
		self.processor_thread.started.connect(self.processor_worker.run)
		self.processor_worker.finished.connect(self.on_processing_finished) # acá recibo los datos
		
		# Limpieza de memoria al terminar
		self.processor_worker.finished.connect(self.processor_thread.quit)
		self.processor_worker.finished.connect(self.processor_worker.deleteLater)
		self.processor_thread.finished.connect(self.processor_thread.deleteLater)

		# Inicio el hilo y muestro el dialog de forma modal
		self.processor_thread.start()
		self.processor_dialog.exec()


	def on_processing_finished(self, final_data):
		"""
		Carga los datos procesados en la UI, actualiza las tablas y muestra
		el reporte final si existe.
		"""

		# Cierro el dialog
		self.processor_dialog.accept()

		# Asigno los datos a la ventana principal
		self.all_products_hh = final_data['hh']['products']
		self.all_products_etma = final_data['etma']['products']
		self.all_products_camba = final_data['camba']['products']
		self.all_products_vtm = final_data['vtm']['products']
		self.report = final_data['report']

		# Mapeo de marcas a sus correspondientes elementos
		bmap = {
			'hh': {
				'products': self.all_products_hh,
				'label': self.label_validity_date_hh,
				'table': self.tableWidget_search_hh,
				'combo': self.comboBox_most_used_hh,
				'most': MOST_USED_PRODUCTS_HH
			},
			'etma': {
				'products': self.all_products_etma,
				'label': self.label_validity_date_etma,
				'table': self.tableWidget_search_etma,
				'combo': self.comboBox_most_used_etma,
				'most': MOST_USED_PRODUCTS_ETMA
			},
			'camba': {
				'products': self.all_products_camba,
				'label': self.label_validity_date_camba,
				'table': self.tableWidget_search_camba,
				'combo': self.comboBox_most_used_camba,
				'most': MOST_USED_PRODUCTS_CAMBA
			},
			'vtm': {
				'products': self.all_products_vtm,
				'label': self.label_validity_date_vtm,
				'table': self.tableWidget_search_vtm,
				'combo': self.comboBox_most_used_vtm,
				'most': MOST_USED_PRODUCTS_VTM
			}
		}

		for brand, elems in bmap.items():
			# Muestro fecha de validez de precios
			elems['label'].setText(final_data[brand]['date'])

			# Listo todos los productos
			self.list_products(elems['products'], elems['table'])

			# Listo los más usados
			self.load_more_used(elems['combo'], elems['products'], elems['most'])

		# Muestro el reporte si existe
		if self.report:
			QMessageBox.information(
				self,
				'Información de la carga',
				self.prepare_report()
			)


	def prepare_report(self):
		"""
		Lee el diccionario de reportes de errores generados por el DataProcessor
		y los formatea en un string amigable para mostrar en un QMessageBox.

		El diccionario de reporte tiene una estructura similar a esta. Solamente
		se agrega algo al diccionario cuando hubo un problema:

		{
			'hh': {
				'excel': {
					'reason': 'no_link',
					'local_status': 'local_used'
				}
			}
			'camba': {
				'excel': {
					'reason': 'no_url',
					'local_status': 'local_used'
				},
				'pdfs': {
					'no_link': {
						'missing': ['05'],
						'local': ['01','04']
					},
					'no_download': {
						'missing': [],
						'local': ['10','11']
					}
				}
			}
		}
		"""
		
		maps = {
			'no_url': 'Sin URL configurada para',
			'no_access': 'Imposible acceder a',
			'no_link': 'No se encontró link',
			'no_download': 'No se pudo descargar',
			'login_failed': 'Fallo de autenticación en',
			'local_used': 'Usando lista local previa',
			'local_missing': 'Lista local no encontrada',
			'local_error': 'Error al procesar lista'
		}
		
		brand_to_supplier = {
			'hh': 'Tienda del Cardan',
			'etma': 'Tienda del Cardan',
			'camba': 'Bulonera Camba',
			'rosario': 'Rosario Agro',
			'vtm': 'VTM Transmisiones'
		}
		
		msg = ''

		# Itero sobre cada marca que tuvo algún problema
		for brand, data in self.report.items():
			# Agrego el título de la marca
			msg += '<br><br>' if msg else ''
			msg += f'<b><u>{brand.upper()}</u></b>'
			
			# PROBLEMAS CON LA LISTA EXCEL DE LA MARCA
			if 'excel' in data:
				# Agrego el tipo de lista
				msg += '<br><b>Lista Excel</b>:'
				
				# Extraigo el estado de la lista local (ej: "local_used")
				local_status = data['excel']['local_status']

				# Extraigo la razón del problema (ej: "no_access")
				# Uso .get() porque si falló al procesar el excel descargado, "reason" no existe
				reason = data['excel'].get('reason')

				# Defino el ícono según si el programa pudo salvar la situación o no
				symbol = '⚠️' if local_status == 'local_used' else '❌'

				# Obtengo el texto que describe el estado local. Ej: "Usando lista local previa"
				local_status_str = maps[local_status]

				# Ajusto el texto si fue un error de procesamiento
				if local_status == 'local_error':
					if reason is None:
						local_status_str += ' recién descargada'
					else:
						local_status_str += ' local'

				# Armo la primera parte de la oración (el motivo del problema)
				if reason:
					reason_str = maps[reason]

					# Si el problema fue de conexión al proveedor, agrego el nombre del mismo
					if reason in ('no_url', 'no_access', 'login_failed'):
						supplier_str = f' <i>{brand_to_supplier[brand]}</i>'
					else:
						supplier_str = ''

					# Ej: " Sin URL configurada para <i>Tienda del Cardan</i>."
					first_part = f' {reason_str}{supplier_str}.'
				else:
					# Si no hay "reason", fue un error directo al procesar, no hay primera parte
					first_part = ''

				# Concateno todo. 
				# Ej 1: " Sin URL configurada para <i>Tienda del Cardan</i>. ⚠️ Usando lista local previa."
				# Ej 2: " ❌ Error al procesar lista recién descargada."
				msg += f'{first_part} {symbol} {local_status_str}.'

			# PROBLEMAS CON LOS ARCHIVOS PDF (CAMBA O ROSARIO)
			if 'pdfs' in data:
				# Agrego el tipo de lista
				msg += '<br><b>Listas PDF</b>:'

				# Itero sobre cada motivo de error (ej: "no_link", "no_download")
				for reason, info in data['pdfs'].items():

					# Junto todos los identificadores de PDFs que fallaron por este motivo
					# Ej: ['01', '02', '05'] o ['Cadenas_LinkBelt', 'Crucetas_ETMA', 'Cuchillas_Agro']
					sheets = info['missing'] + info['local']
					sheets.sort()

					# Ajusto gramática (singular o plural de la palabra "Hoja")
					s = '' if len(sheets) == 1 else 's'

					# Agrego el proveedor si fue un error de conexión a la página del mismo
					supplier_str = f' <i>{brand_to_supplier[brand]}</i>' if reason in ('no_url', 'no_access') else ''

					# Ej: "<br>- Hojas 01, 02: Imposible acceder a <i>Bulonera Camba</i>."
					# Ej: "<br>- Hoja 05: No se pudo descargar."
					msg += f'<br>- Hoja{s} {", ".join(sheets)}: {maps[reason]}{supplier_str}.'

					# PDFs que se pudieron salvar con archivos locales previos
					if info['local']:
						if set(info['local']) == set(sheets) and len(sheets) > 1:
							# Todos los que fallaron tenían respaldo local
							sheets_str = 'todas ellas'
						else:
							# Solo algunos tenían respaldo
							sheets_str = ', '.join(info['local'])

						# Ej: " ⚠️ Usando lista local previa para todas ellas."
						# Ej: " ⚠️ Usando lista local previa para 01."
						msg += f' ⚠️ Usando lista local previa para {sheets_str}.'

					# PDFs que se perdieron completamente (no había local)
					if info['missing']:
						if set(info['missing']) == set(sheets) and len(sheets) > 1:
							# Ninguno de los que fallaron tenía respaldo local
							sheets_str = 'ninguna de ellas'
						else:
							# Faltaron respaldos específicos
							sheets_str = ', '.join(info['missing'])

						# Ej: " ❌ Lista local no encontrada para ninguna de ellas."
						# Ej: " ❌ Lista local no encontrada para 05."
						msg += f' ❌ Lista local no encontrada para {sheets_str}.'

		return msg


	def initialize_theme(self):
		"""Determina y aplica el tema inicial."""

		# Recupero tema guardado (si lo hay)
		saved_theme = SETTINGS.value('theme', '', type=str)

		if saved_theme in ('dark', 'light'):
			initial_theme = saved_theme
		else:
			# Detecto esquema de color actual del sistema
			color_scheme = QApplication.instance().styleHints().colorScheme()
			initial_theme = 'dark' if color_scheme == Qt.ColorScheme.Dark else 'light'

		self.apply_theme(initial_theme)


	def apply_theme(self, theme):
		"""Aplica el tema y cambia los iconos en función del tema."""

		app = QApplication.instance()

		# Guardo en QSettings el tema aplicado
		SETTINGS.setValue('theme', theme)

		# Cambio esquema de color de la app
		if theme == 'dark':
			app.styleHints().setColorScheme(Qt.ColorScheme.Dark)
			self.tableWidget_fans.itemDelegate().set_line_color('#191919')
		else:
			app.styleHints().setColorScheme(Qt.ColorScheme.Light)
			self.tableWidget_fans.itemDelegate().set_line_color('#c8c8c8')

		# Actualizo ícono de botones
		self.pushButton_theme.setIcon(QIcon(f'resources/icons/icon_mode_{theme}.svg'))
		self.pushButton_config.setIcon(QIcon(f'resources/icons/icon_config_{theme}.svg'))
		self.pushButton_about.setIcon(QIcon(f'resources/icons/icon_about_{theme}.svg'))


	def change_theme(self):
		"""Invierte el tema actual."""

		current_theme = SETTINGS.value('theme', type=str)
		new_theme ='light' if current_theme == 'dark' else 'dark'
		self.apply_theme(new_theme)


	def format_headers(self):
		"""Distribuye el ancho de las columnas de todas las tablas."""

		brand_tables = (
			self.tableWidget_search_hh,
			self.tableWidget_defaults_hh,
			self.tableWidget_search_etma,
			self.tableWidget_defaults_etma,
			self.tableWidget_search_camba,
			self.tableWidget_defaults_camba,
			self.tableWidget_search_vtm,
			self.tableWidget_defaults_vtm
		)

		four_cols_tables = (
			self.tableWidget_search_camba,
			self.tableWidget_defaults_camba,
			self.tableWidget_search_vtm,
			self.tableWidget_defaults_vtm
		)

		for table in brand_tables:
			header = table.horizontalHeader()

			table.setColumnWidth(0, 110) # Columna 0 (Código): Fija

			if table in four_cols_tables:
				table.setColumnWidth(1, 400) # Columna 1 (Subcategoría): Fija
				header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch) # Columna 2 (Descripción): Estirada
				table.setColumnWidth(3, 180) # Columna 3 (Precio): Fija
			else:
				header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch) # Columna 1 (Descripción): Estirada
				table.setColumnWidth(2, 180) # Columna 2 (Precio): Fija

		# Tabla de ventiladores
		fans_table = self.tableWidget_fans

		fans_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
		for col in (0, 2, 3, 4):
			fans_table.setColumnWidth(col, 150)


	def init_db(self):
		"""Crea la base de datos local y su contenido si no existen."""

		DB_DIR.mkdir(parents=True, exist_ok=True)

		with sqlite3.connect(DB_PATH) as conn:
			cursor = conn.cursor()
			cursor.execute('''
				CREATE TABLE IF NOT EXISTS fans_data (
					row   INTEGER,
					col   INTEGER,
					value TEXT,
					PRIMARY KEY (row, col)
				)
			''')


	def load_fans_data(self):
		"""Carga los valores guardados al abrir la aplicación."""

		# Bloqueo señales para no activar save_fan_cell()
		self.tableWidget_fans.blockSignals(True)
		
		with sqlite3.connect(DB_PATH) as conn:
			cursor = conn.cursor()
			cursor.execute('SELECT row, col, value FROM fans_data')
			
			for row, col, value in cursor.fetchall():
				item = self.tableWidget_fans.item(row, col)
				if item:
					item.setText(value)
				else:
					# Si la celda estaba vacía y sin item desde Qt Designer, lo creo
					self.tableWidget_fans.setItem(row, col, QTableWidgetItem(value))
					
		self.tableWidget_fans.blockSignals(False)


	def save_fan_cell(self, item):
		"""Guarda automáticamente en la BD cada vez que el usuario edita una celda."""

		row = item.row()
		col = item.column()
		value = item.text().strip()
		
		with sqlite3.connect(DB_PATH) as conn:
			cursor = conn.cursor()
			cursor.execute('''
				INSERT OR REPLACE INTO fans_data (row, col, value)
				VALUES (?, ?, ?)
			''', (row, col, value))


	def load_more_used(self, combo_box, all_products, most_used_products):
		"""Carga los productos más usados."""

		combo_box.setPlaceholderText('Seleccione una categoría...')

		# Detecto si estoy cargando el combo de HH
		is_hh = combo_box is self.comboBox_most_used_hh

		# Cargo categorías y sus productos por detrás
		for category, products_in_category in most_used_products.items():
			products = []
			for product_code, custom_description in products_in_category.items():
				# Busco el producto dentro de todos los productos
				for product in all_products:
					if product_code == product['code']:
						# Si es HH, ignoro la descripción del diccionario y uso la original
						desc_to_use = product['description'] if is_hh else custom_description

						prod_dict = {
							'code': product['code'],
							'description': desc_to_use,
							'price': product['price']
						}

						# Traspaso la subcategoría solo si existe
						if 'subcategory' in product:
							prod_dict['subcategory'] = product['subcategory']

						products.append(prod_dict)

			combo_box.addItem(category, products)

		# Establezco que no haya uno seleccionado
		combo_box.setCurrentIndex(-1)


	def load_category(self):
		"""Lista los productos mas usados de la categoría seleccionada."""

		# Determino si se seleccionó categoría en HH, ETMA, o CAMBA, y asigno variables
		sender = self.sender()
		if sender is self.comboBox_most_used_hh:
			table_widget = self.tableWidget_defaults_hh
			combo_box = self.comboBox_most_used_hh
		elif sender is self.comboBox_most_used_etma:
			table_widget = self.tableWidget_defaults_etma
			combo_box = self.comboBox_most_used_etma
		elif sender is self.comboBox_most_used_camba:
			table_widget = self.tableWidget_defaults_camba
			combo_box = self.comboBox_most_used_camba
		else:
			table_widget = self.tableWidget_defaults_vtm
			combo_box = self.comboBox_most_used_vtm

		# Vacio la tabla y listo los productos
		table_widget.setRowCount(0)
		self.list_products(combo_box.currentData(), table_widget)


	def filter_products(self, query):
		"""Filtra la lista de productos al escribir en el buscador."""

		sender = self.sender()

		# Determino si se buscó en HH o en ETMA, y asigno variables
		if sender is self.lineEdit_search_hh:
			table_widget = self.tableWidget_search_hh
			all_products = self.all_products_hh
		elif sender is self.lineEdit_search_etma:
			table_widget = self.tableWidget_search_etma
			all_products = self.all_products_etma
		elif sender is self.lineEdit_search_camba:
			table_widget = self.tableWidget_search_camba
			all_products = self.all_products_camba
		else:
			table_widget = self.tableWidget_search_vtm
			all_products = self.all_products_vtm

		# Evito lógica innecesaria si no se cargaron productos en la marca
		if not all_products:
			return

		# Divido el filtro en palabras
		query_words = query.lower().split()

		# Busco productos coincidentes
		if query_words:
			filtered_products = []
			for product in all_products:
				match = True
				for word in query_words:
					# Uso .get('subcategory', '') por si no existe la clave
					if (word not in product['code'].lower()
						and word not in product.get('subcategory', '').lower() 
						and word not in product['description'].lower()
					):
						match = False
				if match:
					filtered_products.append(product)
			self.list_products(filtered_products, table_widget)
		else: # Si no hay nada escrito, muestro todos los productos
			self.list_products(all_products, table_widget)


	def list_products(self, products, table_widget):
		"""Lista los productos en la tabla correspondiente."""

		table_widget.setRowCount(0)

		# Detecto si la tabla tiene 4 columnas
		is_4_cols = table_widget in (
				self.tableWidget_search_camba, 
				self.tableWidget_defaults_camba, 
				self.tableWidget_search_vtm, 
				self.tableWidget_defaults_vtm)

		for product in products:
			row = table_widget.rowCount()
			table_widget.insertRow(row)

			# Columna 0: Siempre el código
			code_item = QTableWidgetItem(product['code'])
			table_widget.setItem(row, 0, code_item)

			# Si es Camba o VTM, mapeo 4 columnas. Si no, mapeo 3.
			if is_4_cols:
				subcat_item = QTableWidgetItem(product.get('subcategory', ''))
				table_widget.setItem(row, 1, subcat_item)

				descr_item = QTableWidgetItem(product['description'])
				table_widget.setItem(row, 2, descr_item)

				price_item = QTableWidgetItem(product['price'])
				price_item.setFont(QFont('Consolas', 12))
				table_widget.setItem(row, 3, price_item)
			else:
				descr_item = QTableWidgetItem(product['description'])
				table_widget.setItem(row, 1, descr_item) # Pasa a ser columna 1

				price_item = QTableWidgetItem(product['price'])
				price_item.setFont(QFont('Consolas', 12))
				table_widget.setItem(row, 2, price_item) # Pasa a ser columna 2

		# Muestro el número de productos listado
		search_tables = {
			self.tableWidget_search_hh: self.label_search_hh,
			self.tableWidget_search_etma: self.label_search_etma,
			self.tableWidget_search_camba: self.label_search_camba,
			self.tableWidget_search_vtm: self.label_search_vtm,
			self.tableWidget_defaults_hh: self.label_most_used_hh,
			self.tableWidget_defaults_etma: self.label_most_used_etma,
			self.tableWidget_defaults_camba: self.label_most_used_camba,
			self.tableWidget_defaults_vtm: self.label_most_used_vtm,
		}
		quantity = len(products)
		s = '' if quantity == 1 else 's'
		search_tables[table_widget].setText(f'{quantity} producto{s} encontrado{s}')


	def empty_everything(self):
		"""Vacia la interfaz para la recarga de listas."""

		# Junto widgets que usan clear()
		widgets = {
			self.lineEdit_search_hh,
			self.lineEdit_search_etma,
			self.lineEdit_search_camba,
			self.lineEdit_search_vtm,
			self.label_search_hh,
			self.label_search_etma,
			self.label_search_camba,
			self.label_search_vtm,
			self.label_most_used_hh,
			self.label_most_used_etma,
			self.label_most_used_camba,
			self.label_most_used_vtm,
			self.label_validity_date_hh,
			self.label_validity_date_etma,
			self.label_validity_date_camba,
			self.label_validity_date_vtm,
			self.comboBox_most_used_hh,
			self.comboBox_most_used_etma,
			self.comboBox_most_used_camba,
			self.comboBox_most_used_vtm
		}

		tables = (
			self.tableWidget_search_hh,
			self.tableWidget_defaults_hh,
			self.tableWidget_search_etma,
			self.tableWidget_defaults_etma,
			self.tableWidget_search_camba,
			self.tableWidget_defaults_camba,
			self.tableWidget_search_vtm,
			self.tableWidget_defaults_vtm
		)

		for widget in widgets:
			widget.clear() 

		for table in tables:
			table.setRowCount(0)


	def open_pdf(self, brand, identifier, page_number=1):
		"""
		Busca el PDF correspondiente y lo abre, en orden de disponibilidad, con:
		* Navegador predeterminado, en la página indicada.
		* Visor PDF predeterminado del sistema, sin poder indicar la página.
		
		Parámetros:
		- brand: 'camba' o 'rosario'
		- identifier: número de hoja ('02') para Camba, o nombre ('Cuchillas_Jardin') para Rosario.
		- page_number: La página donde se quiere arrancar (por defecto 1).
		"""

		base_path = LISTS_DIR / brand
		pdf_file_path = None

		if not base_path.exists():
			supplier = 'Bulonera Camba' if brand == 'camba' else 'Rosario Agro'
			QMessageBox.warning(
				self, 
				'Carpeta no encontrada', 
				f'No existe la carpeta de listas para {supplier}.'
			)
			return

		# Busco la ruta del PDF respetando mayúsculas/minúsculas
		if brand == 'camba':
			pdf_files = list(base_path.glob(f'Hoja{identifier}*.pdf'))
			if pdf_files:
				pdf_file_path = pdf_files[0]
		elif brand == 'rosario':
			exact_path = base_path / f'{identifier}.pdf'
			if exact_path.exists():
				pdf_file_path = exact_path

		# # Busco la ruta del PDF ignorando mayúsculas/minúsculas
		# if brand == 'camba':
		# 	target_prefix = f'hoja{identifier}'.lower()
		# 	for pdf_file in base_path.glob('*.pdf'):
		# 		if pdf_file.name.lower().startswith(target_prefix):
		# 			pdf_file_path = pdf_file
		# 			break  # Encontré el archivo, salgo del ciclo
					
		# elif brand == 'rosario':
		# 	target_name = f'{identifier}.pdf'.lower()
		# 	for pdf_file in base_path.glob('*.pdf'):
		# 		if pdf_file.name.lower() == target_name:
		# 			pdf_file_path = pdf_file
		# 			break  # Encontré el archivo, salgo del ciclo

		# Si encontré el archivo, lo abro
		if pdf_file_path:
			try:
				# Intento con el navegador predeterminado
				default_browser_exe = get_default_browser_exe()
				if default_browser_exe:
					# Formateo la ruta de Windows a un formato URI que el navegador entienda
					pdf_uri = f'file:///{str(pdf_file_path).replace(os.sep, "/")}#page={page_number}'
					subprocess.Popen([default_browser_exe, pdf_uri])
				else:
					# Como último recurso, intento con el lector de PDF predeterminado
					url = QUrl.fromLocalFile(str(pdf_file_path))
					QDesktopServices.openUrl(url)
			except Exception as e:
				QMessageBox.critical(
					self, 
					'Error', 
					f'No se pudo abrir el PDF:\n{str(e)}'
				)
		else:
			filename = f'Hoja {identifier}' if brand == 'camba' else identifier
			QMessageBox.warning(
				self, 
				'Archivo no encontrado', 
				f'No se pudo encontrar el PDF local para: <b>{filename}</b>.'
			)


	def open_tdc_catalog(self, item):
		"""
		Abre el catálogo web forzando el inicio de sesión previo para obtener precios
		con descuento.
		"""
		
		# 1. Extraigo el SKU de la tabla
		table = self.sender() 
		row = item.row()
		sku_item = table.item(row, 0)
		
		if not sku_item:
			return
		sku = sku_item.text()

		# 2. Obtengo credenciales y el navegador predeterminado
		email = os.getenv('USER_EMAIL')
		password = os.getenv('USER_PASSWORD')
		default_browser = get_default_browser_exe()

		# Función interna que ejecutará Playwright en segundo plano
		def run_playwright():
			from playwright.sync_api import sync_playwright

			with sync_playwright() as p:
				user_data_dir = APP_DATA_DIR / 'BrowserProfile'
				
				is_new_browser = False
				
				try:
					# 1. Intento conectarme a un navegador que ya esté abierto
					browser = p.chromium.connect_over_cdp('http://localhost:9222')
					context = browser.contexts[0]
					page = context.new_page()
				except Exception:
					# 2. Si falla, lo lanzo de cero
					launch_args = {
						'user_data_dir': user_data_dir,
						'headless': False,
						'args': ['--start-maximized', '--remote-debugging-port=9222'],
						'no_viewport': True
					}

					if default_browser:
						launch_args['executable_path'] = default_browser
					else:
						launch_args['channel'] = 'msedge' 

					try:
						context = p.chromium.launch_persistent_context(**launch_args)
						
						# 1. Obligo a crear una pestaña nueva y segura para el producto actual
						page = context.new_page()
						is_new_browser = True
						
						# 2. Limpieza: cierro la pestaña "about:blank" inicial 
						for tab in context.pages:
							if tab != page and tab.url == 'about:blank':
								try:
									tab.close()
								except:
									pass
					except Exception as e:
						print(f'Error al lanzar el navegador: {e}')
						return
				
				# --- PASO 1: IR DIRECTAMENTE AL PRODUCTO ---
				product_url = f'https://app.tiendadecardan.com.ar/catalogo/{sku}'
				page.goto(product_url)

				try:
					# Espero a que la página cargue para que el botón exista en el HTML
					page.wait_for_load_state('networkidle', timeout=5000)
				except:
					pass

				# --- PASO 2: VERIFICAR SI FALTA LOGUEO ---
				# Uso .first para evitar el error de "strict mode" si hay varios botones de login
				login_btn = page.locator('a[href="/login"]').first
				
				if login_btn.is_visible():
					try:
						# Hago clic en el botón para ir a la pantalla de login
						login_btn.click()
						
						# Relleno el formulario
						page.wait_for_selector('input[type="email"]', timeout=5000)
						page.fill('input[type="email"]', email)
						page.fill('input[type="password"]', password)
						page.press('input[type="password"]', 'Enter')
						
						# Esperamos la redirección automática que hace la página
						page.wait_for_url('**/catalogo*', timeout=15000)
						
						# Vuelvo a cargar el producto, ahora con el descuento aplicado
						page.goto(product_url)
					except Exception as e:
						print(f'Error en el proceso de auto-login: {e}')

				# --- FINALIZAR ---
				if is_new_browser:
					try:
						page.wait_for_event('close', timeout=0)
					except:
						pass

		# Ejecuto Playwright en un hilo separado
		threading.Thread(target=run_playwright, daemon=True).start()


	def open_vtm_catalog(self, item):
		"""Abre el catálogo web de VTM usando Playwright en el NAVEGADOR PREDETERMINADO."""
		
		# 1. Extraigo el SKU de la tabla (igual que para TDC)
		table = self.sender() 
		row = item.row()
		sku_item = table.item(row, 0)
		
		if not sku_item:
			return
		sku = sku_item.text()

		# 2. Obtengo la ruta del ejecutable del navegador PREDETERMINADO del sistema
		default_browser_exe = get_default_browser_exe()

		# Función interna que ejecutará Playwright en segundo plano
		def run_playwright():
			from playwright.sync_api import sync_playwright

			with sync_playwright() as p:
				# Carpeta de perfil persistente (aislada para evitar bloqueos)
				user_data_dir = APP_DATA_DIR / 'BrowserProfile'
				
				# Bandera para saber si lanzamos el navegador o abrimos pestaña
				is_new_browser = False
				
				try:
					# 1. Intento conectarme a un navegador que ya esté abierto
					browser = p.chromium.connect_over_cdp('http://localhost:9222')
					context = browser.contexts[0]
					page = context.new_page()
					
				except Exception:
					# 2. Si falla, lo lanzo de cero
					launch_args = {
						'user_data_dir': user_data_dir,
						'headless': False,
						'args': ['--start-maximized', '--remote-debugging-port=9222'],
						'no_viewport': True
					}

					if default_browser_exe:
						launch_args['executable_path'] = default_browser_exe
					else:
						launch_args['channel'] = 'msedge'
					
					try:
						context = p.chromium.launch_persistent_context(**launch_args)
						
						# 1. Obligo a crear una pestaña nueva y segura para el producto actual
						page = context.new_page()
						is_new_browser = True
						
						# 2. Limpieza: cierro la pestaña "about:blank" inicial 
						for tab in context.pages:
							if tab != page and tab.url == 'about:blank':
								try:
									tab.close()
								except:
									pass
					except Exception as e:
						print(f'Error al lanzar el navegador: {e}')
						return
				
				# --- PASO 1: NAVEGAR A LA PÁGINA DEL CATÁLOGO ---
				# Dado que VTM no tiene URLs directas por SKU y no requiere login,
				# voy a la raíz de la lista.
				page.goto('https://vtm-lista.pages.dev/')

				# --- PASO 2: BUSCAR EL PRODUCTO ---
				try:
					# Espero a que el buscador (identificado por su clase CSS) esté visible
					search_input = page.locator('input.toolbar-search-input')
					search_input.wait_for(state='visible', timeout=10000)
					
					# Relleno el SKU del producto
					search_input.fill(sku)
					
					# Presiono "Enter" para buscar
					page.press('input.toolbar-search-input', 'Enter')
					
					# IMPORTANTE: Mantenemos vivo el hilo principal del navegador si lo lanzamos de cero.
					if is_new_browser:
						page.wait_for_event('close', timeout=0)
						
				except Exception as e:
					print(f"Error al buscar el producto en VTM: {e}")
					# Mantenemos vivo el hilo si lo lanzamos de cero
					if is_new_browser:
						page.wait_for_event('close', timeout=0)

		# Ejecuto Playwright en un hilo separado
		threading.Thread(target=run_playwright, daemon=True).start()


	def open_config(self):
		"""Abre un dialog para editar la configuración."""

		dialog = ConfigurationDialog(self)
		dialog.exec()

		# Verifico si recargar
		if dialog.new_supplier_url:
			self.start_data_processing()


	def open_about(self):
		"""Abre un dialog de Acerca de."""

		dialog = AboutDialog(self)
		dialog.exec()



class ConfigurationDialog(QDialog):
	def __init__(self, parent=None):
		super().__init__(parent)

		# Cargo la UI
		uic.loadUi('ui/config.ui', self)

		self.load_config()

		# Flag para recargar al cerrar dialog
		self.new_supplier_url = False

		# Conecto señales
		self.pushButton_ok.clicked.connect(self.save_config)
		self.pushButton_cancel.clicked.connect(self.close)


	def load_config(self):
		self.lineEdit_url_camba.setText(SETTINGS.value('supplier_urls/camba', '', type=str))


	def save_config(self):
		SETTINGS.setValue('supplier_urls/camba', self.lineEdit_url_camba.text())
		self.new_supplier_url = True # Para recargar al cerrar configuración
		self.close()



class AboutDialog(QDialog):
	def __init__(self, parent=None):
		super().__init__(parent)

		# Cargo la UI
		uic.loadUi('ui/about.ui', self)



class ProgressDialog(QDialog):
	def __init__(self, title, cancellable=True, parent=None):
		super().__init__(parent)

		# Cargo la UI
		uic.loadUi('ui/progress.ui', self)

		self.setWindowTitle(title)
		self.cancellable = cancellable
		self.pushButton_cancel.clicked.connect(self.reject)

		if not self.cancellable:
			# Deshabilito la 'X' de la ventana
			self.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)

			# Escondo el botón de Cancelar
			self.pushButton_cancel.hide()

			# Ajusto y fijo la altura del dialog manteniendo el ancho
			current_width = self.width()
			self.adjustSize()
			self.setFixedSize(current_width, self.height())
		else:
			# Fijo el tamaño original que vino de Qt Designer (con el botón visible)
			self.setFixedSize(self.width(), self.height())


	def reject(self):
		"""
		Atrapa el botón Cancelar, la tecla Escape y la 'X'.
		Si no es cancelable, ignora la orden de cierre.
		"""
		if not self.cancellable:
			return 

		# Si ES cancelable, ejecuta el cierre normal
		super().reject()



# Inicializo la app
if __name__ == "__main__":
	app = QApplication(sys.argv)

	# Establezco tema de aplicación
	app.setStyle('Fusion')

	# Configuro traducción al español de botones
	translator = QTranslator()
	path = QLibraryInfo.path(QLibraryInfo.LibraryPath.TranslationsPath)
	if translator.load('qtbase_es', path):
		app.installTranslator(translator)

	window = MainWindow()
	window.show()
	sys.exit(app.exec())